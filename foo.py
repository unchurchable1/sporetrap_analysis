import openpyxl
from openpyxl.styles import Font

# Create a dictionary to map values to font colors
value_color_mapping = {
    "1": "FF0000",  # Red
    "2": "00FF00",  # Green
    "3": "0000FF",  # Blue
    "4": "C0C0C0",  # Silver
    "5": "FFD700"   # Gold
}

# Load the Excel file
file_path = "SporeTrap_Workbook.xlsx"  # Replace with the path to your Excel file
workbook = openpyxl.load_workbook(file_path)

# Iterate through each sheet in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    # Find the column index of "Position" (assuming it's in the first row)
    position_column_index = None
    for col_index, cell in enumerate(sheet[1], start=1):
        if cell.value == "Position":
            position_column_index = col_index
            break
    
    # Check if "Position" column was found
    if position_column_index is not None:
        # Iterate through rows starting from the second row (assuming headers are in the first row)
        for row in sheet.iter_rows(min_row=2, min_col=position_column_index, max_col=position_column_index):
            for cell in row:
                cell_value = str(cell.value)  # Convert cell value to string
                if cell_value in value_color_mapping:
                    # Change the font color based on the value
                    font_color = value_color_mapping[cell_value]
                    cell.font = Font(color=font_color)

# Save the modified workbook in-place
workbook.save(file_path)

# Close the workbook
workbook.close()

print("Excel file modified in-place.")
