#!/usr/bin/env python3
#
# This file is part of the sporetrap analysis scripts.
#
# Copyright (c) 2024 Jason Toney
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""Combines multiple CSV results files into a single Excel workbook."""

import os
import csv

import openpyxl


def compile_workbook(workbook_file, csv_files):
    """
    Compiles the csv files into the workbook.
    """
    # Check if output workbook already exists
    if os.path.exists(workbook_file):
        # If it does, load the existing workbook
        workbook = openpyxl.load_workbook(workbook_file)
    else:
        # Otherwise, create a new workbook
        workbook = openpyxl.Workbook()

        # Remove the default sheet created by openpyxl
        default_sheet = workbook["Sheet"]
        workbook.remove(default_sheet)

    # Add each new sheet to the workbook if it does not already exist
    for file in csv_files:
        # Create a new sheet in the workbook
        sheet_name = os.path.splitext(os.path.basename(file))[0]
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=sheet_name)

            # Open the csv file and read in the data
            with open(file, "r", encoding="utf-8") as csv_file:
                csv_reader = csv.reader(csv_file)

                # Loop through the rows  and add them to the sheet
                for row in csv_reader:
                    sheet.append(row)

            print(f"Added sheet {sheet_name} to workbook")
        else:
            print(f"Skipped sheet {sheet_name}: already in the workbook")

    # Save the workbook
    workbook.save(workbook_file)

    # Log the number of sheets in the workbook
    print(f"The workbook contains {len(workbook.sheetnames)} sheets")


def main():
    """Set path and get filenames"""
    # Change to the script's directory
    os.chdir(os.path.dirname(__file__))

    # output file
    workbook_file = "Particle Release Counts - CST filters.xlsx"

    # Get list of all usable csv files in the results directory
    csv_files = []
    for file in os.listdir("results"):
        if not file.endswith(".csv"):
            continue
        file_path = os.path.join("results", file)
        csv_files.append(file_path)

    if csv_files:
        compile_workbook(workbook_file, csv_files)
    else:
        print("No csv files found.")


if __name__ == "__main__":
    main()
