#!/usr/bin/env python3
#
# This file is part of the sporetrap analysis scripts.
#
# Copyright (c) 2024 Jason Toney
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""
    docstring
"""

import csv
import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment


def add_filters(sheet):
    """docstring"""
    # Iterate through rows starting from the second row
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            # Use actual filter sizes instead of relative positions
            cell_value = int(cell.value)
            positions = ["Foil", "50 µm Mesh", "150 µm Mesh", "400 µm Mesh"]
            cell.value = positions[cell_value - 1]


def add_manual_counts(counts_file, sheet):
    """docstring"""
    # Load up the counts
    matching_values = []
    with open(counts_file, "r", encoding="utf-8", newline="") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            matching_values.append(row)

    # Iterate through rows
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=2):
        trap_value = row[0].value
        position_value = str(row[1].value)
        # Add notations if available
        for match_data in matching_values:
            if (
                match_data["Trap"] == trap_value
                and match_data["Position"] == position_value
            ):
                sheet.cell(row=row[0].row, column=4, value=match_data["Microspheres"])


def autosize_columns(sheet):
    """docstring"""
    for column in sheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(
            column[0].column
        )  # Get column letter
        for cell in column:
            try:  # Avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width


def format_workbook(filename):
    """docstring"""
    # Load the Excel file
    workbook = openpyxl.load_workbook(filename)

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Add manually counted Red microspheres to the workbook
        counts_file = f"{os.path.dirname(__file__)}/counts/{sheet_name} - Red.csv"
        if os.path.exists(counts_file):
            print(f"Adding manually counted microspheres to sheet {sheet_name}")
            add_manual_counts(counts_file, sheet)

        # Rename the position column with filter sizes
        add_filters(sheet)

        # Auto-size columns to fit content
        autosize_columns(sheet)

    # Save the modified workbook in-place
    workbook.save(filename)
    # Close the workbook
    workbook.close()

    print(f"Workbook: {os.path.basename(filename)} has been reformatted.")


def main(filename):
    """Execute main objective."""
    os.chdir(os.path.dirname(__file__))
    format_workbook(filename)


if __name__ == "__main__":
    if len(sys.argv) == 2:
        main(sys.argv[1])
    else:
        sys.exit(f"Usage: python3 {__file__} [FILE]")
